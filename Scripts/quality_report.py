import pandas as pd 

tabela = pd.read_excel("Dados/Checklist QA Fic.xlsx")

#print(tabela)
#dado = tabela.loc[tabela["Responsável"]=="Aegon VI Targaryen"]

# Do exemplo do vídeo tabela.loc[tabela["Tipo"]=="Serviço", "Multiplicador"] = 1.5
# Na coluna Tipo, nas linhas que possuem o item Serviço, altera a coluna Multiplicador para 1.5
# Não altera as outras linhas da tabela, só as que filtram certo como Serviço (não alterou as linhas com Produto, por exemplo)
#print(dado)
# def item1_1():
#    #dado = tabela.loc[tabela["Nome ponto de decisão"].isin(["Execução", "Exec. Projeto"])] -----Funciona, filtra duas informações em uma coluna
#    #dado = tabela.loc[tabela["Nome ponto de decisão"].isin(["Execução", "Exec. Projeto"]) & tabela["Data real de início"].isnull()]
#    dado = tabela.loc[tabela["Nome ponto de decisão"].isin(["Execução", "Exec. Projeto"]) & tabela["Data real de início"].isnull() | tabela["Dt.real in.pto.dec"].isnull()]
#    print(dado)
   
# item1_1()

# def item1_2():
#    dado = tabela.loc[tabela["Nome ponto de decisão"].isin(["Exec. Business Case"]) & tabela["Dt.real in.pto.dec"].isnull()]
#    print(dado)

# item1_2()

# def item2():
#    filt1 = tabela["Nome ponto de decisão"].isin(["Exec. Projeto", "Exec. Business Case"]) 
#    filt2 = tabela["Status ponto de decisão"].isin(["Planejamento Exec.", "Ag. Execução"])

#    cond1 = tabela["Dt. Plan início"] <= pd.to_datetime("2024-02-14") 
#    cond2 = tabela["Data real de início"].isnull()
#    cond3 = tabela["Dt. Plan início"].isnull()
#    cond4 = tabela["Data real de início"].notnull()

#    print("Projetos atrasados ainda não iniciados")
#    print(tabela.loc[filt1 & filt2 & cond1 & cond2][["ID", "Nome", "Responsável", "Nome ponto de decisão", "Status ponto de decisão", "Dt. Plan início"]])

#    print("Projetos sem data de início planejada")
#    print(tabela.loc[filt1 & filt2 & cond3][["ID", "Nome", "Responsável", "Nome ponto de decisão", "Status ponto de decisão"]])

#    print("Projetos com status incorreto")
#    print(tabela.loc[filt1 & filt2 & cond4][["ID", "Nome", "Responsável", "Nome ponto de decisão", "Status ponto de decisão", "Data real de início"]])

# item2()

#colunas = ["ID", "Nome", "Responsável", "Nome ponto de decisão", "Status ponto de decisão", "Dt. Plan início", "Data real de início"]
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment

def item2():
    # Definindo as colunas que serão impressas
    colunas = ["ID", "Nome", "Responsável", "Nome ponto de decisão", "Status ponto de decisão", "Dt. Plan início", "Data real de início"]

    filt1 = tabela["Nome ponto de decisão"].isin(["Exec. Projeto", "Exec. Business Case"])
    filt2 = tabela["Status ponto de decisão"].isin(["Planejamento Exec.", "Ag. Execução"])

    cond1 = tabela["Dt. Plan início"] <= pd.to_datetime("2024-02-14")
    cond2 = tabela["Data real de início"].isnull()
    cond3 = tabela["Dt. Plan início"].isnull()
    cond4 = tabela["Data real de início"].notnull()

    # Criando a pasta de trabalho
    wb = Workbook()

    # Criando uma planilha para cada condição
    planilha_atrasados = wb.create_sheet("Projetos Atrasados")
    planilha_sem_data = wb.create_sheet("Projetos Sem Data")
    planilha_status_incorreto = wb.create_sheet("Projetos Status Incorreto")

    # Definindo o estilo das células
    estilo_titulo = Alignment(horizontal="center", vertical="center")

    # Adicionando os títulos das colunas
    for i, coluna in enumerate(colunas):
        planilha_atrasados.cell(row=1, column=i+1).value = coluna
        planilha_sem_data.cell(row=1, column=i+1).value = coluna
        planilha_status_incorreto.cell(row=1, column=i+1).value = coluna

    # Contador para controlar a linha atual
    linha_atual = 2

    # Percorrendo as linhas da tabela original
    for i in range(len(tabela)):
        # Verificando as condições
        if filt1[i] & filt2[i]:
            if cond1[i] & cond2[i]:
                # Adicionando os dados à planilha de projetos atrasados
                j = 1
                for coluna in colunas:
                    planilha_atrasados.cell(row=linha_atual, column=j).value = tabela[coluna].iloc[i]
                    j += 1
                linha_atual += 1
            elif cond3[i]:
                # Adicionando os dados à planilha de projetos sem data
                j = 1
                for coluna in colunas:
                    planilha_sem_data.cell(row=linha_atual, column=j).value = tabela[coluna].iloc[i]
                    j += 1
                linha_atual += 1
            elif cond4[i]:
                # Adicionando os dados à planilha de projetos com status incorreto
                j = 1
                for coluna in colunas:
                    planilha_status_incorreto.cell(row=linha_atual, column=j).value = tabela[coluna].iloc[i]
                    j += 1
                linha_atual += 1

    # Autofiltrando as planilhas para remover linhas vazias
    planilha_atrasados.auto_filter.ref = "A1:D" + str(linha_atual - 1)
    planilha_sem_data.auto_filter.ref = "A1:D" + str(linha_atual - 1)
    planilha_status_incorreto.auto_filter.ref = "A1:D" + str(linha_atual - 1)

    planilha_atrasados.auto_filter.add_filter_column(1, ["(Blanks)"])
    planilha_sem_data.auto_filter.add_filter_column(1, ["(Blanks)"])
    planilha_status_incorreto.auto_filter.add_filter_column(1, ["(Blanks)"])

    # Salvando o arquivo Excel
    wb.save("exibição_tabelas.xlsx")

    # Abrindo o arquivo Excel
    os.startfile("exibição_tabelas.xlsx")

item2()


   
# colunas = tabela.columns 
# print(colunas)
